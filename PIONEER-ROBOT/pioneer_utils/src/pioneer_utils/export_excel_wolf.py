#!/usr/bin/env python3

import os
import rospy
import rospkg
import xlsxwriter
import numpy as np 
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.create_file()
        
    def create_file(self):
        if not os.path.exists(self.file_path):
            workbook  = xlsxwriter.Workbook(self.file_path)
            worksheet = workbook.add_worksheet()

            header_format = workbook.add_format({
                # 'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': 'center'})

            worksheet.set_column(0, 0, 5)   # Column  A width set to 5
            worksheet.set_column('B:O', 15) # set column B:J width
            # worksheet.set_row(0, 30)        # set row 0 height to 40

            worksheet.write('A1', 'No.',          header_format)
            worksheet.write('B1', 'imu_roll',     header_format)
            worksheet.write('C1', 'imu_pitch',    header_format)
            worksheet.write('D1', 'imu_yaw',      header_format)
            worksheet.write('E1', 'left_foot_x',  header_format)
            worksheet.write('F1', 'left_foot_y',  header_format)
            worksheet.write('G1', 'left_foot_z',  header_format)
            worksheet.write('H1', 'right_foot_x', header_format)
            worksheet.write('I1', 'right_foot_y', header_format)
            worksheet.write('J1', 'right_foot_z', header_format)
            worksheet.write('K1', 'des_roll',     header_format)
            worksheet.write('L1', 'des_pitch',    header_format)
            worksheet.write('M1', 'robot_frame',  header_format)
            worksheet.write('N1', 'tripod_frame', header_format)
            worksheet.write('O1', 'rs_frame',     header_format)
            
        else:
            rospy.loginfo('[Excel] File : {} is exist'.format(self.file_path))

    def read_data(self):
        reader = pd.read_excel(self.file_path)
        print(reader)

    def add_data(self, no, imu_roll, imu_pitch, imu_yaw, \
                           lf_x, lf_y, lf_z, \
                           rf_x, rf_y, rf_z, \
                           des_roll, des_pitch, \
                           robot_frame, tripod_frame, rs_frame ):

        workbook  = load_workbook(self.file_path)
        worksheet = workbook.active

        new_row = [[no, imu_roll, imu_pitch, imu_yaw, \
                        lf_x,     lf_y,      lf_z,    \
                        rf_x,     rf_y,      rf_z,    \
                        des_roll, des_pitch,          \
                        robot_frame, tripod_frame, rs_frame]]

        for data in new_row:
            worksheet.append(data)

        max_row    = worksheet.max_row
        max_column = worksheet.max_column

        for j in range(1, max_column+1):
            cell = worksheet.cell(row=max_row,column=j)
            cell.alignment = Alignment(horizontal='center') 

        workbook.save(filename=self.file_path)