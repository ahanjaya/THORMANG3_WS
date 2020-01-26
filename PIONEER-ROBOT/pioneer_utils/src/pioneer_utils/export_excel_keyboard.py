#!/usr/bin/env python3

import os
import rospy
import rospkg
import xlsxwriter
import numpy as np 
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.create_file()
        
    def create_file(self):
        if not os.path.exists(self.file_path):
            workbook  = xlsxwriter.Workbook(self.file_path)
            worksheet = workbook.add_worksheet()

            header_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': 'center'})

            worksheet.set_column(0, 0, 5)   # Column  A width set to 5
            worksheet.set_column('B:O', 13) # set column width
            worksheet.set_row(0, 30)        # set row 0 height to 40

            worksheet.write('A1', 'No.',                      header_format)
            worksheet.write('B1', 'Actual Start (X)',         header_format)
            worksheet.write('C1', 'Actual Start (Y)',         header_format)
            worksheet.write('D1', 'Actual Start (Theta)',     header_format)
            worksheet.write('E1', 'Actual Final (X)',         header_format)
            worksheet.write('F1', 'Actual Final (Y)',         header_format)
            worksheet.write('G1', 'Actual Final (Theta)',     header_format)
            worksheet.write('H1', 'Simulation Final (X)',     header_format)
            worksheet.write('I1', 'Simulation Final (Y)',     header_format)
            worksheet.write('J1', 'Simulation Final (Theta)', header_format)
            worksheet.write('K1', 'Position Error (Pixel)',   header_format)
            worksheet.write('L1', 'Position Error (Cm)',      header_format)
            worksheet.write('M1', 'Theta Error',              header_format)
            worksheet.write('N1', 'Mapping Theta Error',      header_format)
            worksheet.write('O1', 'Status',                   header_format)
            
        else:
            rospy.loginfo('[Excel] File : {} is exist'.format(self.file_path))

    def read_data(self):
        reader = pd.read_excel(self.file_path)
        print(reader)

    def add_data(self, no, x_start_actual, y_start_actual, theta_start_actual, \
                           x_final_actual, y_final_actual, theta_final_actual, \
                           x_simula, y_simula, theta_simula, \
                           pos_err_pixel,  pos_err_cm, theta_err, map_theta_err, status ):

        workbook  = load_workbook(self.file_path)
        worksheet = workbook.active

        new_row = [[no, x_start_actual, y_start_actual, theta_start_actual, \
                        x_final_actual, y_final_actual, theta_final_actual, \
                        x_simula, y_simula, theta_simula, \
                        pos_err_pixel, pos_err_cm, theta_err, map_theta_err, status ]]

        for data in new_row:
            worksheet.append(data)

        max_row    = worksheet.max_row
        max_column = worksheet.max_column

        for j in range(1, max_column+1):
            cell = worksheet.cell(row=max_row,column=j)
            cell.alignment = Alignment(horizontal='center') 

        # # print all value
        # for i in range(1,max_row+1):
        #     for j in range(1,max_column+1):
        #         cell = worksheet.cell(row=i,column=j)
        #         print(cell.value,end=' | ')
        #     print('\n')

        workbook.save(filename=self.file_path)

#     def run(self):
#         self.read_data()
#         self.add_data (no=1, x_actual=2, y_actual=3, theta_actual=4, \
#                              x_simula=5, y_simula=6, theta_simula=7, \
#                              pos_err=8,  theta_err=9, map_theta_err=10 )

# if __name__ == '__main__':
#     rospack   = rospkg.RosPack()
#     file_path = rospack.get_path("pioneer_main") + "/data/"

#     excel = Excel(file_path + 'keyboard_placement.xlsx')
#     # excel.run()